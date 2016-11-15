#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 14:59:44 2016

@author:Beiwen Liu
"""

import pandas as pd
import requests
from openpyxl import Workbook
import openpyxl
import os.path
from django.utils.encoding import smart_str
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl import Workbook

def main():
    excelFile = raw_input("Please choose an excel file by name:\n")
    excel(excelFile)
    
def excel(name):
    wb = openpyxl.load_workbook('{}.xlsx'.format(name), data_only=True) #Commands
    my_file = '{}Result.xlsx'.format(name) #Results
    if ~os.path.isfile(my_file):
        createExcel(name, my_file) #If results excel doesnt exist, create it
    wb2 = openpyxl.load_workbook(my_file)
    sheet_names = wb.get_sheet_names()
    for s in sheet_names:
        sheet = wb.get_sheet_by_name(s)
        sheet2 = wb2.get_sheet_by_name(s)
        traverseSheet(musicSheet(sheet,sheet2),sheet,sheet2)
    wb2.save(my_file)
    
#Create result excel file. Copies sheets
def createExcel(name, my_file):
    new_file = my_file
    print "executed"
    wb1 = openpyxl.load_workbook('{}.xlsx'.format(name))
    sheet_names = wb1.get_sheet_names()
    wb = openpyxl.Workbook()
    counter = 0
    for x in sheet_names:
        wb.create_sheet(index=counter, title=x) #Creating sheets based on original excel
        counter = counter + 1
    wb.save(new_file)

def request(text):
    features ={"text": text,"seat":1,"demo_id":"d6608a6456af027e82cf14737ddc7e09"}
    r = requests.post('https://mono-v.mybluemix.net/conversation', data=features)
    return r.json()
    
        
def musicSheet(sheet,sheet2):
    sheet2['A1'] = sheet['A1'].value
    if str(sheet['A1'].value) == "Command":
        command = 'A'
    sheet2['A1'].font = Font(bold=True, size=18)
    sheet2['B1'] = sheet['B1'].value
    if str(sheet['B1'].value) == "Command":
        command = 'B'
    sheet2['B1'].font = Font(bold=True, size=18)
    sheet2['C1'] = sheet['C1'].value
    sheet2['C1'].font = Font(bold=True, size=18)
    if str(sheet['C1'].value) == "Command":
        command = 'C'
    sheet2['D1'] = "Response"
    sheet2['D1'].font = Font(bold=True, size=18)
    if str(sheet['D1'].value) == "Command":
        command = 'D'
    sheet2['E1'] = "Action Type"
    sheet2['E1'].font = Font(bold=True, size=18)
    if str(sheet['E1'].value) == "Command":
        command = 'E'
            
    return command        
def traverseSheet(command,sheet,sheet2):
    row_num = sheet.max_row
    for x in range(2,row_num + 1):
        tempStr = smart_str(sheet[command + str(x)].value)
        resp = request(smart_str(sheet[command + str(x)].value))
        print resp['text']
        
        #if len(resp['actions']) == 0:
        sheet2['A' + str(x)] = smart_str(sheet['A' + str(x)].value)
        sheet2['B' + str(x)] = smart_str(sheet['B' + str(x)].value)
        sheet2['C' + str(x)] = smart_str(sheet['C' + str(x)].value)
        sheet2['D' + str(x)] = smart_str(resp['text'])
        if len(resp['actions']) != 0:
            sheet2['E' + str(x)] = smart_str(resp['actions'][0]['action_type'])
            sheet2['F' + str(x)] = smart_str(resp['actions'])
            if smart_str(resp['actions'][0]['action_type']) != smart_str(sheet['B' + str(x)].value):
                sheet2['A' + str(x)].font = Font(color=colors.RED)
                sheet2['B' + str(x)].font = Font(color=colors.RED)
                sheet2['C' + str(x)].font = Font(color=colors.RED)
                sheet2['D' + str(x)].font = Font(color=colors.RED)
                sheet2['E' + str(x)].font = Font(color=colors.RED)
                sheet2['F' + str(x)].font = Font(color=colors.RED)
                sheet2['G' + str(x)] = "Action type mismatch"
                sheet2['G' + str(x)].font = Font(color=colors.RED)
        else:
            sheet2['A' + str(x)].font = Font(color=colors.RED)
            sheet2['B' + str(x)].font = Font(color=colors.RED)
            sheet2['C' + str(x)].font = Font(color=colors.RED)
            sheet2['D' + str(x)].font = Font(color=colors.RED)
            sheet2['E' + str(x)] = "Missing Action"
            sheet2['E' + str(x)].font = Font(color=colors.RED)
        
        
        
main()
"""
def s():
    a = request("play adele")
    print a['text']
    print a['actions']
    for e in a['actions'][0]:
        print e
        print a['actions'][0][e]

s()"""
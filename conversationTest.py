#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 10 14:59:44 2016

@author: MacbookRetina
"""

import pandas as pd
import requests
from openpyxl import Workbook
import openpyxl
import os.path


def createExcel():
    new_file = 'flowResult.xlsx'
    print "executed"
    wb1 = openpyxl.load_workbook('flow.xlsx')
    sheet_names = wb1.get_sheet_names()
    wb = openpyxl.Workbook()
    counter = 0
    for x in sheet_names:
        wb.create_sheet(index=counter, title=x)
        counter = counter + 1
    wb.save(new_file)

def request(text):
    features ={"text": text,"seat":1,"demo_id":"d6608a6456af027e82cf14737ddc7e09"}
    r = requests.post('https://mono-v.mybluemix.net/conversation', data=features)
    return str(r.json()['text'])
    
def excel():
    wb = openpyxl.load_workbook('flow.xlsx')
    my_file = 'flowResult.xlsx'
    if ~os.path.isfile(my_file):
        createExcel()
    wb2 = openpyxl.load_workbook('flowResult.xlsx')
    sheet_names = wb.get_sheet_names()
    for s in sheet_names:
        sheet = wb.get_sheet_by_name(s)
        sheet2 = wb2.get_sheet_by_name(s)
        traverseSheet(sheet,sheet2)
    wb2.save(my_file)
        
def traverseSheet(sheet,sheet2):
    row_num = sheet.max_row
    for x in range(2,row_num + 1):
        resp = request(str(sheet['A' + str(x)].value))
        answer = resp == str(sheet['B' + str(x)].value)
        print resp
        sheet2['A' + str(x)] = str(sheet['A' + str(x)].value)
        sheet2['B' + str(x)] = str(sheet['B' + str(x)].value)
        if (answer):
            sheet2['C' + str(x)] = "Same"
        else:
            sheet2['C' + str(x)] = "Not same"
            sheet2['D' + str(x)] = resp

#main("hello")

excel()

#createExcel()